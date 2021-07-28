import requests
from bs4 import BeautifulSoup
import re
import json

headers = {
    "Accept": "*/*",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/70.0.3538.77 Safari/537.36 "
}

# делаем запрос на сайт для получения html кода которрый имеет нуджную нам инфорацию
# req = requests.get('https://rskrf.ru/ratings/tekhnika-i-elektronika/electronic/smartfony-/', headers=headers)
# src = req.text


# сохраняем в html на случай если сайт пошлет нах за парсинг(много папрсинг запросов и сайт может блокнуть)
# with open("index.html", "w", encoding='utf-8') as file:
#     file.write(src)

try:
    with open("index.html", "r", encoding="utf-8") as file:
        src = file.read()

    soup = BeautifulSoup(src, "lxml")

    # название модели
    a = r"^\S+"
    # тип модели + номер если есть
    b = r"- (.*)\s\("
    b1 = r"\(.*\)|\s-\s.*"
    # гигабайт
    c = r"(\d+GB)"
    # наличие 5G (да | нет)
    d = r"5G"
    # шкала качества
    f = r"\S[\w+]\S+\d+\S+$"
    model_site_info = {}

    # поиск текста  по тегу(ссылке)
    for i in soup.find_all(href=re.compile(r"goods+")):
        # сылка на модель
        # model_href = ("href = https://rskrf.ru" + i.get(r"href"))
        model_href = (r"https://rskrf.ru" + i.get(r"href"))
        print(model_href)

        # название модели
        model_phone_name = (re.match(a, i.text))
        print(model_phone_name.group(0))

        # тип модели
        model_phone = (re.findall(b, i.text))
        for iteem in model_phone:
            model_phone = iteem
        model_phone = re.sub(b1, "", model_phone)
        model_phone = re.sub('5G', "", model_phone)
        print(model_phone)

        # проверка на наличие 5G
        test_5g = (re.findall(d, i.text))
        if len(test_5g) != 0:
            test_5g = 'Yes'
            print(test_5g)
        else:
            test_5g = "None 5G"
            print(test_5g)

        # проверка наличия гигабайт
        memory = (re.findall(c, i.text))
        if len(memory) != 0:
            repository = memory[0]
            print(repository)
        else:
            repository = "None GB"
            print(repository)

        # последний элемент - средня шкала качества
        general_quality_scale = (re.findall(f, i.text))
        print(general_quality_scale[0], "\n")

        # добавляем и охраняем все данные в формате словаря.
        model_site_info[model_phone] = {'href': model_href,
                                        'Model': model_phone_name.group(0),
                                        'Access_5G': test_5g,
                                        'Internal_Storage': repository,
                                        'Average_Rating': general_quality_scale[0]}

    # print(model_site_info)

    # добаввляю данные из списка в json.
    with open('models.json', 'w') as file:
        json.dump(model_site_info, file, indent=4)

    # заполняем полученными данными из спика в таблицу exel
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    row = 0
    for model_phone in model_site_info:
        model_stat = []
        # это костыль
        for val in model_site_info[model_phone].values():
            model_stat.append(val)
        row += 1
        ws[f'A{row}'] = model_phone
        ws[f'B{row}'] = model_stat[1]
        ws[f'C{row}'] = model_stat[2]
        ws[f'D{row}'] = model_stat[3]
        ws[f'E{row}'] = model_stat[4]
        ws[f'F{row}'] = model_stat[0]
    wb.save("example.xlsx")
except FileNotFoundError:
    print('Произошел сбой программы. нету необходимого html документа для работы, попробуй попробовать еще раз.')
    # делаем запрос на сайт для получения html кода которрый имеет нуджную нам инфорацию
    req = requests.get('https://rskrf.ru/ratings/tekhnika-i-elektronika/electronic/smartfony-/', headers=headers)
    src = req.text


    # сохраняем в html на случай если сайт пошлет нах за парсинг(много папрсинг запросов и сайт может блокнуть)
    with open("index.html", "w", encoding='utf-8') as file:
        file.write(src)